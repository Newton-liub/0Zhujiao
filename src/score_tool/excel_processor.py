from __future__ import annotations

import re
from statistics import median
from collections import Counter, OrderedDict, defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import (
    REVIEW_RESULT_MODIFIED,
    MergeData,
    ProcessResult,
    ReviewDecision,
    ReviewExportResult,
    ReviewFlag,
    ReviewSessionItem,
    SourceFileConfig,
    SourcePreview,
    StudentScoreRow,
)

REQUIRED_HEADERS = {
    "student_id": ("学号/工号", "学号", "工号"),
    "student_name": ("学生姓名", "姓名"),
    "class_name": ("班级",),
    "score": ("总分",),
}

MAX_HEADER_SCAN_ROWS = 20


@dataclass(frozen=True)
class ParsedClass:
    raw: str
    prefix: str
    grade: int | None
    major: int | None
    class_no: int | None
    tail: str


@dataclass(frozen=True)
class ScoreScale:
    name: str
    full_mark: float
    low_threshold: float


@dataclass(frozen=True)
class CellReview:
    level: str
    reasons: tuple[str, ...]


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).replace("\u3000", " ").strip()


def normalize_header(value: Any) -> str:
    return re.sub(r"\s+", "", normalize_text(value))


def normalize_student_id(value: Any) -> str:
    text = normalize_text(value)
    if re.fullmatch(r"\d+\.0", text):
        return text[:-2]
    return text


def parse_score(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize_text(value)
    if not text:
        return None
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    return float(match.group(0)) if match else None


def infer_score_scale(values: Iterable[Any]) -> ScoreScale | None:
    numeric_values = [score for value in values if (score := parse_score(value)) is not None]
    if not numeric_values:
        return None

    max_score = max(numeric_values)
    if max_score <= 10:
        return ScoreScale(name="10分制", full_mark=10.0, low_threshold=8.0)
    if max_score <= 100:
        return ScoreScale(name="百分制", full_mark=100.0, low_threshold=60.0)

    return ScoreScale(name=f"{max_score:g}分制", full_mark=max_score, low_threshold=max_score * 0.6)


def normalize_to_percent(score: float | None, scale: ScoreScale | None) -> float | None:
    if score is None or scale is None or scale.full_mark <= 0:
        return None
    return score / scale.full_mark * 100


def is_close_score(score: float, target: float) -> bool:
    return abs(score - target) <= 1e-6


def is_blank_row(values: Iterable[Any]) -> bool:
    return all(normalize_text(value) == "" for value in values)


def detect_header(row_values: list[Any]) -> dict[str, int] | None:
    headers = [normalize_header(value) for value in row_values]
    found: dict[str, int] = {}
    for key, candidates in REQUIRED_HEADERS.items():
        for index, header in enumerate(headers):
            if header in candidates:
                found[key] = index
                break
        if key not in found:
            return None
    return found


def infer_score_name(path: Path) -> str:
    stem = path.stem.strip()
    patterns = [
        r"(阶段\s*测试\s*\d+)",
        r"(阶段\s*测验\s*\d+)",
        r"(测试\s*\d+)",
        r"(测验\s*\d+)",
    ]
    for pattern in patterns:
        match = re.search(pattern, stem)
        if match:
            return re.sub(r"\s+", "", match.group(1))

    separators = ["-", "_", "—", "–"]
    for separator in separators:
        if separator in stem:
            candidate = stem.split(separator)[-1].strip()
            if candidate:
                return candidate
    return stem or "总分"


def unique_names(names: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for name in names:
        base = name or "总分"
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def parse_class_name(class_name: str) -> ParsedClass:
    text = normalize_text(class_name)
    match = re.search(r"^(?P<prefix>\D*?)(?P<grade>\d{2})(?P<major>\d{2})-(?P<class_no>\d+)(?P<tail>.*)$", text)
    if match:
        return ParsedClass(
            raw=text,
            prefix=match.group("prefix"),
            grade=int(match.group("grade")),
            major=int(match.group("major")),
            class_no=int(match.group("class_no")),
            tail=match.group("tail"),
        )

    loose = re.search(r"(?P<grade>\d{2})", text)
    return ParsedClass(
        raw=text,
        prefix=re.sub(r"\d.*$", "", text),
        grade=int(loose.group("grade")) if loose else None,
        major=None,
        class_no=None,
        tail="",
    )


def natural_student_id_key(student_id: str) -> tuple[int, Any]:
    text = normalize_student_id(student_id)
    if re.fullmatch(r"\d+", text):
        return (0, int(text))
    return (1, text)


def class_sort_key(class_name: str) -> tuple[Any, ...]:
    parsed = parse_class_name(class_name)
    return (
        parsed.prefix,
        parsed.grade if parsed.grade is not None else 999,
        parsed.major if parsed.major is not None else 999,
        parsed.class_no if parsed.class_no is not None else 999,
        parsed.tail,
        parsed.raw,
    )


def detect_main_grade(class_names: Iterable[str]) -> int | None:
    grades = [parse_class_name(name).grade for name in class_names]
    known_grades = [grade for grade in grades if grade is not None]
    if not known_grades:
        return None
    return Counter(known_grades).most_common(1)[0][0]


def is_retake(class_name: str, main_grade: int | None) -> bool:
    if main_grade is None:
        return False
    grade = parse_class_name(class_name).grade
    return grade is not None and grade != main_grade


def read_source(config: SourceFileConfig) -> tuple[SourcePreview, list[StudentScoreRow]]:
    path = config.path
    warnings: list[str] = []
    workbook = load_workbook(path, data_only=True, read_only=True)

    selected_sheet = None
    header_row_number = 0
    header_map: dict[str, int] | None = None

    for worksheet in workbook.worksheets:
        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_row=min(MAX_HEADER_SCAN_ROWS, worksheet.max_row), values_only=True),
            start=1,
        ):
            detected = detect_header(list(row))
            if detected is not None:
                selected_sheet = worksheet
                header_row_number = row_index
                header_map = detected
                break
        if selected_sheet is not None:
            break

    if selected_sheet is None or header_map is None:
        raise ValueError(f"{path.name} 未找到必需表头：学号/工号、学生姓名、班级、总分")

    rows: list[StudentScoreRow] = []
    seen_ids: set[str] = set()
    max_col = selected_sheet.max_column
    data_start = header_row_number + 1

    for row in selected_sheet.iter_rows(min_row=data_start, max_col=max_col, values_only=True):
        if is_blank_row(row):
            continue

        student_id = normalize_student_id(row[header_map["student_id"]] if header_map["student_id"] < len(row) else None)
        student_name = normalize_text(row[header_map["student_name"]] if header_map["student_name"] < len(row) else None)
        class_name = normalize_text(row[header_map["class_name"]] if header_map["class_name"] < len(row) else None)
        score = row[header_map["score"]] if header_map["score"] < len(row) else None

        if not student_id:
            continue
        if student_id in seen_ids:
            warnings.append(f"{path.name} 中学号/工号 {student_id} 重复，已保留最后一次成绩")
            rows = [item for item in rows if item.student_id != student_id]
        seen_ids.add(student_id)

        if not student_name:
            warnings.append(f"{path.name} 中学号/工号 {student_id} 缺少学生姓名")
        if not class_name:
            warnings.append(f"{path.name} 中学号/工号 {student_id} 缺少班级")
        if score is None or normalize_text(score) == "":
            warnings.append(f"{path.name} 中学号/工号 {student_id} 总分为空")

        rows.append(
            StudentScoreRow(
                student_id=student_id,
                student_name=student_name,
                class_name=class_name,
                score=score,
                source_path=path,
            )
        )

    score_name = normalize_text(config.score_name) or infer_score_name(path)
    preview = SourcePreview(
        path=path,
        sheet_name=selected_sheet.title,
        header_row=header_row_number,
        student_count=len(rows),
        score_name=score_name,
        warnings=warnings,
    )
    return preview, rows


def preview_sources(configs: list[SourceFileConfig]) -> list[SourcePreview]:
    previews: list[SourcePreview] = []
    raw_names: list[str] = []
    for config in configs:
        preview, _ = read_source(config)
        previews.append(preview)
        raw_names.append(preview.score_name)

    for preview, score_name in zip(previews, unique_names(raw_names), strict=True):
        preview.score_name = score_name
    return previews


def analyze_review_flags(
    students: list[dict[str, Any]],
    score_columns: list[str],
) -> tuple[list[ReviewFlag], dict[tuple[str, str], CellReview]]:
    scales = {
        score_column: infer_score_scale(student["scores"].get(score_column) for student in students)
        for score_column in score_columns
    }
    score_values = {
        student["学号/工号"]: {
            score_column: parse_score(student["scores"].get(score_column))
            for score_column in score_columns
        }
        for student in students
    }
    normalized_values = {
        student_id: {
            score_column: normalize_to_percent(score, scales[score_column])
            for score_column, score in scores.items()
        }
        for student_id, scores in score_values.items()
    }

    flags: list[ReviewFlag] = []
    seen: set[tuple[str, str, str]] = set()

    def add_flag(
        level: str,
        student: dict[str, Any],
        score_column: str,
        reason: str,
        suggestion: str,
    ) -> None:
        student_id = student["学号/工号"]
        key = (student_id, score_column, reason)
        if key in seen:
            return
        seen.add(key)
        flags.append(
            ReviewFlag(
                level=level,
                student_id=student_id,
                student_name=student["学生姓名"],
                class_name=student["班级"],
                score_column=score_column,
                score=student["scores"].get(score_column, ""),
                reason=reason,
                suggestion=suggestion,
            )
        )

    for score_column in score_columns:
        scale = scales[score_column]
        if scale is None:
            continue
        for student in students:
            score = score_values[student["学号/工号"]][score_column]
            if score is None:
                continue
            if is_close_score(score, scale.full_mark):
                add_flag(
                    "核查",
                    student,
                    score_column,
                    f"{scale.name}满分 {scale.full_mark:g}，请确认是否确为满分。",
                    "核对原始成绩、平台导出记录和登分表是否一致。",
                )
            if score <= scale.low_threshold:
                add_flag(
                    "核查",
                    student,
                    score_column,
                    f"{scale.name}成绩 {score:g} 小于等于核查线 {scale.low_threshold:g}。",
                    "核对原始成绩，确认是否需要保及格或其他人工处理。",
                )

    def find_swap_candidate(
        index: int,
        student: dict[str, Any],
        score_column: str,
        current_percent: float,
        own_baseline: float,
    ) -> tuple[str, dict[str, Any]] | None:
        for offset, label in ((-1, "上一位"), (1, "下一位")):
            neighbor_index = index + offset
            if neighbor_index < 0 or neighbor_index >= len(students):
                continue
            neighbor = students[neighbor_index]
            if neighbor["班级"] != student["班级"]:
                continue

            neighbor_id = neighbor["学号/工号"]
            neighbor_current = normalized_values[neighbor_id].get(score_column)
            neighbor_others = [
                value
                for column, value in normalized_values[neighbor_id].items()
                if column != score_column and value is not None
            ]
            if neighbor_current is None or len(neighbor_others) < 2:
                continue

            neighbor_baseline = median(neighbor_others)
            before = abs(current_percent - own_baseline) + abs(neighbor_current - neighbor_baseline)
            after = abs(neighbor_current - own_baseline) + abs(current_percent - neighbor_baseline)
            if after + 15 <= before and abs(neighbor_current - own_baseline) <= 15 and abs(current_percent - neighbor_baseline) <= 15:
                return label, neighbor
        return None

    for index, student in enumerate(students):
        student_id = student["学号/工号"]
        normalized_scores = normalized_values[student_id]
        for score_column, current_percent in normalized_scores.items():
            if current_percent is None:
                continue
            other_scores = [
                value
                for column, value in normalized_scores.items()
                if column != score_column and value is not None
            ]
            if len(other_scores) < 2:
                continue

            own_baseline = median(other_scores)
            if own_baseline - current_percent < 20 or current_percent > min(other_scores) - 10:
                continue

            swap_candidate = find_swap_candidate(index, student, score_column, current_percent, own_baseline)
            if swap_candidate is not None:
                label, neighbor = swap_candidate
                add_flag(
                    "重点核查",
                    student,
                    score_column,
                    f"该次成绩明显低于本人其他成绩；与{label}同学 {neighbor['学号/工号']} 对调后更接近双方平时水平。",
                    "优先核对原表相邻两行、导出顺序、复制粘贴和手工录入记录。",
                )
            else:
                add_flag(
                    "核查",
                    student,
                    score_column,
                    "该次成绩明显低于本人其他成绩。",
                    "核对该学生该次原始成绩，确认是否漏录、错列或导入错位。",
                )

    cell_flags: dict[tuple[str, str], list[ReviewFlag]] = defaultdict(list)
    for flag in flags:
        cell_flags[(flag.student_id, flag.score_column)].append(flag)

    cell_reviews = {
        key: CellReview(
            level="重点核查" if any(flag.level == "重点核查" for flag in grouped_flags) else "核查",
            reasons=tuple(flag.reason for flag in grouped_flags),
        )
        for key, grouped_flags in cell_flags.items()
    }
    return flags, cell_reviews


def format_review_summary(flags: list[ReviewFlag]) -> str:
    if not flags:
        return ""
    focus = [flag for flag in flags if flag.level == "重点核查"]
    normal = [flag for flag in flags if flag.level != "重点核查"]
    parts: list[str] = []
    if focus:
        parts.append("重点核查：" + "；".join(f"{flag.score_column}（{flag.reason}）" for flag in focus))
    if normal:
        parts.append("核查：" + "；".join(f"{flag.score_column}（{flag.reason}）" for flag in normal))
    return "；".join(parts)


def build_cell_reviews(review_flags: list[ReviewFlag]) -> dict[tuple[str, str], CellReview]:
    cell_flags: dict[tuple[str, str], list[ReviewFlag]] = defaultdict(list)
    for flag in review_flags:
        cell_flags[(flag.student_id, flag.score_column)].append(flag)

    return {
        key: CellReview(
            level="重点核查" if any(flag.level == "重点核查" for flag in grouped_flags) else "核查",
            reasons=tuple(flag.reason for flag in grouped_flags),
        )
        for key, grouped_flags in cell_flags.items()
    }


def build_merge_data(configs: list[SourceFileConfig]) -> MergeData:
    if not configs:
        raise ValueError("请至少选择一个 Excel 文件")

    previews: list[SourcePreview] = []
    rows_by_source: list[list[StudentScoreRow]] = []
    raw_score_names: list[str] = []

    for config in configs:
        preview, rows = read_source(config)
        previews.append(preview)
        rows_by_source.append(rows)
        raw_score_names.append(preview.score_name)

    score_columns = unique_names(raw_score_names)
    for preview, score_name in zip(previews, score_columns, strict=True):
        preview.score_name = score_name

    students: OrderedDict[str, dict[str, Any]] = OrderedDict()
    warnings: list[str] = []
    for preview in previews:
        warnings.extend(preview.warnings)

    for source_index, rows in enumerate(rows_by_source):
        score_column = score_columns[source_index]
        for row in rows:
            current = students.setdefault(
                row.student_id,
                {
                    "学号/工号": row.student_id,
                    "学生姓名": row.student_name,
                    "班级": row.class_name,
                    "scores": {},
                },
            )

            if current["学生姓名"] and row.student_name and current["学生姓名"] != row.student_name:
                warnings.append(f"学号/工号 {row.student_id} 姓名不一致：{current['学生姓名']} / {row.student_name}")
            if current["班级"] and row.class_name and current["班级"] != row.class_name:
                warnings.append(f"学号/工号 {row.student_id} 班级不一致：{current['班级']} / {row.class_name}")

            if not current["学生姓名"] and row.student_name:
                current["学生姓名"] = row.student_name
            if not current["班级"] and row.class_name:
                current["班级"] = row.class_name
            current["scores"][score_column] = row.score

    main_grade = detect_main_grade(student["班级"] for student in students.values())
    for student in students.values():
        if parse_class_name(student["班级"]).grade is None:
            warnings.append(f"学号/工号 {student['学号/工号']} 的班级无法识别年级：{student['班级']}")

    ordered_students = sorted(
        students.values(),
        key=lambda student: (
            0 if is_retake(student["班级"], main_grade) else 1,
            class_sort_key(student["班级"]),
            natural_student_id_key(student["学号/工号"]),
        ),
    )
    review_flags, _ = analyze_review_flags(ordered_students, score_columns)
    return MergeData(
        source_previews=previews,
        students=ordered_students,
        score_columns=score_columns,
        warnings=warnings,
        main_grade=main_grade,
        review_flags=review_flags,
    )


def merge_sources(configs: list[SourceFileConfig], output_path: Path, include_log_sheet: bool = True) -> ProcessResult:
    merge_data = build_merge_data(configs)
    cell_reviews = build_cell_reviews(merge_data.review_flags)
    focus_review_count = sum(1 for flag in merge_data.review_flags if flag.level == "重点核查")

    write_output(
        output_path,
        merge_data.students,
        merge_data.score_columns,
        merge_data.source_previews,
        merge_data.warnings,
        include_log_sheet,
        merge_data.main_grade,
        merge_data.review_flags,
        cell_reviews,
    )
    return ProcessResult(
        output_path=output_path,
        source_previews=merge_data.source_previews,
        row_count=len(merge_data.students),
        score_columns=merge_data.score_columns,
        review_count=len(merge_data.review_flags),
        focus_review_count=focus_review_count,
        merge_data=merge_data,
        warnings=merge_data.warnings,
    )


def write_output(
    output_path: Path,
    students: list[dict[str, Any]],
    score_columns: list[str],
    previews: list[SourcePreview],
    warnings: list[str],
    include_log_sheet: bool,
    main_grade: int | None,
    review_flags: list[ReviewFlag],
    cell_reviews: dict[tuple[str, str], CellReview],
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "成绩汇总"

    review_flags_by_student: dict[str, list[ReviewFlag]] = defaultdict(list)
    for flag in review_flags:
        review_flags_by_student[flag.student_id].append(flag)

    headers = ["学号/工号", "学生姓名", "班级", *score_columns, "核查标记"]
    worksheet.append(headers)
    score_column_indexes = {score_column: headers.index(score_column) + 1 for score_column in score_columns}
    review_column_index = len(headers)
    focus_fill = PatternFill("solid", fgColor="F4CCCC")
    review_fill = PatternFill("solid", fgColor="FFF2CC")
    summary_fill = PatternFill("solid", fgColor="FCE5CD")

    for row_index, student in enumerate(students, start=2):
        student_id = student["学号/工号"]
        summary = format_review_summary(review_flags_by_student.get(student_id, []))
        worksheet.append(
            [
                student_id,
                student["学生姓名"],
                student["班级"],
                *[student["scores"].get(score_column, "") for score_column in score_columns],
                summary,
            ]
        )
        if summary:
            worksheet.cell(row=row_index, column=review_column_index).fill = summary_fill

        for score_column in score_columns:
            review = cell_reviews.get((student_id, score_column))
            if review is None:
                continue
            cell = worksheet.cell(row=row_index, column=score_column_indexes[score_column])
            cell.fill = focus_fill if review.level == "重点核查" else review_fill
            cell.comment = Comment("\n".join(review.reasons), "ScoreTool")

    style_worksheet(worksheet)

    if review_flags:
        review_sheet = workbook.create_sheet("核查明细")
        review_sheet.append(["级别", "学号/工号", "学生姓名", "班级", "成绩列", "成绩", "原因", "建议"])
        for flag in sorted(review_flags, key=lambda item: (0 if item.level == "重点核查" else 1, item.class_name, item.student_id, item.score_column)):
            review_sheet.append([
                flag.level,
                flag.student_id,
                flag.student_name,
                flag.class_name,
                flag.score_column,
                flag.score,
                flag.reason,
                flag.suggestion,
            ])
        for row in review_sheet.iter_rows(min_row=2):
            row[0].fill = focus_fill if row[0].value == "重点核查" else review_fill
        style_worksheet(review_sheet)

    if include_log_sheet:
        log_sheet = workbook.create_sheet("处理日志")
        log_sheet.append(["项目", "内容"])
        log_sheet.append(["主年级", main_grade if main_grade is not None else "未识别"])
        log_sheet.append(["输出人数", len(students)])
        log_sheet.append(["核查标记", f"{len(review_flags)} 条，其中重点核查 {sum(1 for flag in review_flags if flag.level == '重点核查')} 条"])
        log_sheet.append([])
        log_sheet.append(["核查规则", "说明"])
        log_sheet.append(["10分制", "标记 10 分和小于等于 8 分；同时检查某次成绩是否明显低于本人其他成绩。"])
        log_sheet.append(["百分制", "标记 100 分和小于等于 60 分；同时检查某次成绩是否明显低于本人其他成绩。"])
        log_sheet.append(["疑似对调", "若异常低分与前后同班同学对调后更接近双方平时水平，则标为重点核查。"])
        log_sheet.append([])
        log_sheet.append(["源文件", "工作表", "表头行", "人数", "成绩列名"])
        for preview in previews:
            log_sheet.append([preview.path.name, preview.sheet_name, preview.header_row, preview.student_count, preview.score_name])
        log_sheet.append([])
        log_sheet.append(["警告"])
        for warning in warnings:
            log_sheet.append([warning])
        style_worksheet(log_sheet)

    workbook.save(output_path)


def make_review_item_key(student_id: str, score_column: str) -> str:
    return f"{student_id}\u241f{score_column}"


def build_review_session_items(merge_data: MergeData) -> list[ReviewSessionItem]:
    flags_by_cell: dict[tuple[str, str], list[ReviewFlag]] = defaultdict(list)
    for flag in merge_data.review_flags:
        flags_by_cell[(flag.student_id, flag.score_column)].append(flag)

    items: list[ReviewSessionItem] = []
    for student in merge_data.students:
        student_id = student["学号/工号"]
        for score_column in merge_data.score_columns:
            flags = flags_by_cell.get((student_id, score_column), [])
            if not flags:
                continue
            items.append(
                ReviewSessionItem(
                    key=make_review_item_key(student_id, score_column),
                    index=len(items),
                    level="重点核查" if any(flag.level == "重点核查" for flag in flags) else "核查",
                    student_id=student_id,
                    student_name=student["学生姓名"],
                    class_name=student["班级"],
                    score_column=score_column,
                    score=student["scores"].get(score_column, ""),
                    reasons=tuple(flag.reason for flag in flags),
                )
            )
    return items


def derive_output_path(base_path: Path, suffix: str) -> Path:
    return base_path.with_name(f"{base_path.stem}{suffix}{base_path.suffix or '.xlsx'}")


def clone_students_with_decisions(
    students: list[dict[str, Any]],
    decisions: dict[str, ReviewDecision],
) -> list[dict[str, Any]]:
    cloned: list[dict[str, Any]] = []
    for student in students:
        copied = {**student, "scores": dict(student["scores"])}
        for decision in decisions.values():
            if decision.result != REVIEW_RESULT_MODIFIED:
                continue
            for score_column in copied["scores"]:
                if decision.item_key == make_review_item_key(copied["学号/工号"], score_column):
                    copied["scores"][score_column] = decision.corrected_score
        cloned.append(copied)
    return cloned


def write_review_report(
    output_path: Path,
    merge_data: MergeData,
    decisions: dict[str, ReviewDecision],
    include_modified: bool,
) -> None:
    items = build_review_session_items(merge_data)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "核查报告"

    headers = [
        "学号/工号",
        "学生姓名",
        "班级",
        *merge_data.score_columns,
        "核查成绩列",
        "原成绩",
        "核查原因",
        "核查结果",
        "修正后成绩",
    ]
    worksheet.append(headers)
    students_by_id = {student["学号/工号"]: student for student in merge_data.students}
    for item in items:
        decision = decisions.get(item.key)
        result = decision.result if decision else "未核查"
        if not include_modified and result == REVIEW_RESULT_MODIFIED:
            continue
        student = students_by_id[item.student_id]
        worksheet.append(
            [
                item.student_id,
                item.student_name,
                item.class_name,
                *[student["scores"].get(score_column, "") for score_column in merge_data.score_columns],
                item.score_column,
                item.score,
                "；".join(item.reasons),
                result,
                decision.corrected_score if decision and decision.corrected_score is not None else "",
            ]
        )
    style_worksheet(worksheet)
    workbook.save(output_path)


def export_review_outputs(
    merge_data: MergeData,
    base_output_path: Path,
    decisions: dict[str, ReviewDecision],
    include_filtered_report: bool,
) -> ReviewExportResult:
    report_path = derive_output_path(base_output_path, "_核查报告")
    write_review_report(report_path, merge_data, decisions, include_modified=True)

    has_modified = any(decision.result == REVIEW_RESULT_MODIFIED for decision in decisions.values())
    corrected_output_path = None
    if has_modified:
        corrected_output_path = derive_output_path(base_output_path, "_已修正")
        corrected_students = clone_students_with_decisions(merge_data.students, decisions)
        write_output(
            corrected_output_path,
            corrected_students,
            merge_data.score_columns,
            merge_data.source_previews,
            merge_data.warnings,
            True,
            merge_data.main_grade,
            merge_data.review_flags,
            build_cell_reviews(merge_data.review_flags),
        )

    filtered_report_path = None
    if include_filtered_report and has_modified:
        filtered_report_path = derive_output_path(base_output_path, "_核查报告_不含登记错误")
        write_review_report(filtered_report_path, merge_data, decisions, include_modified=False)

    return ReviewExportResult(
        report_path=report_path,
        corrected_output_path=corrected_output_path,
        filtered_report_path=filtered_report_path,
    )


def style_worksheet(worksheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    worksheet.freeze_panes = "A2"
    for column_cells in worksheet.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
        for cell in column_cells:
            value = normalize_text(cell.value)
            max_length = max(max_length, len(value))
        worksheet.column_dimensions[column_letter].width = min(max(max_length + 2, 10), 40)


def build_configs(paths: list[str | Path], score_names: dict[str, str] | None = None) -> list[SourceFileConfig]:
    score_names = score_names or {}
    configs: list[SourceFileConfig] = []
    for item in paths:
        path = Path(item)
        configs.append(SourceFileConfig(path=path, score_name=score_names.get(str(path))))
    return configs